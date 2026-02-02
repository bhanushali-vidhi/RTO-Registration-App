import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
import pytesseract
from pdf2image import convert_from_bytes
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- PAGE CONFIG ---
st.set_page_config(page_title="RTO Document Verifier + OCR", layout="wide")

# --- HELPER FUNCTIONS ---

def normalize_text(text):
    if not text: return ""
    text = re.sub(r'[^\w\s]', ' ', str(text))
    return text.lower().strip()

def check_name_match(excel_name, doc_name):
    if not doc_name or not excel_name:
        return False
    
    clean_excel = normalize_text(excel_name)
    clean_doc = normalize_text(doc_name)

    excel_tokens = clean_excel.split()
    doc_tokens = clean_doc.split()

    matches = 0
    for doc_word in doc_tokens:
        if doc_word in excel_tokens:
            matches += 1
            continue
        if len(doc_word) == 1:
            if any(token.startswith(doc_word) for token in excel_tokens):
                matches += 1
                continue

    if len(doc_tokens) > 0 and (matches / len(doc_tokens)) >= 0.5:
        return True
    return False

def extract_text_from_pdf_upload(uploaded_file):
    text_content = ""
    try:
        # 1. Try Digital Extraction first
        with pdfplumber.open(uploaded_file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text: text_content += text + "\n"
        
        # 2. Check if extraction was successful or if it's a scanned image
        # If text length is very short, it's likely a scan.
        if len(text_content.strip()) < 20:
            uploaded_file.seek(0)  # Reset file pointer
            # Convert PDF to images (one per page)
            images = convert_from_bytes(uploaded_file.read())
            
            ocr_text_list = []
            for img in images:
                # Perform OCR on the image
                ocr_text_list.append(pytesseract.image_to_string(img))
            
            text_content = "\n".join(ocr_text_list)
            
    except Exception as e:
        st.error(f"Error processing file: {e}")
        return ""
    return text_content

def parse_document_data(text):
    data = {}
    low_text = text.lower()
    
    # 1. GLOBAL TEMPORARY CHECK
    if "temporary" in low_text:
        data['reg_type'] = "Temporary"
        data['vehicle_no'] = "TEMP"
    else:
        data['reg_type'] = "Unknown" 

    # 2. Find Chassis No
    chassis_match = re.search(r'\b[A-HJ-NPR-Z0-9]{17}\b', text)
    data['doc_chassis'] = chassis_match.group(0) if chassis_match else None

    # 3. Find Vehicle No
    if data['reg_type'] != "Temporary":
        perm_pattern = r'\b[A-Z]{2}[0-9]{1,2}[A-Z]{1,3}[0-9]{4}\b'
        bh_pattern = r'\b[0-9]{2}BH[0-9]{4}[A-Z]{1,2}\b'
        
        found_perm = re.search(perm_pattern, text)
        found_bh = re.search(bh_pattern, text)

        if found_perm or found_bh:
            data['reg_type'] = "Permanent"
            data['vehicle_no'] = (found_perm or found_bh).group(0)
        else:
            if "new" in low_text:
                data['vehicle_no'] = "NEW"
                data['reg_type'] = "Temporary"
            else:
                data['vehicle_no'] = "Not Found"
                data['reg_type'] = "Temporary" 

    # 4. Find Customer Name
    name_match = re.search(r'(?:Received From|Customer Name|Name|Mr\.|Ms\.)[:\s\.]*([A-Za-z\s\.]+)', text, re.IGNORECASE)
    if name_match:
        raw_name = name_match.group(1).strip()
        data['doc_name'] = " ".join(raw_name.split()[:4]) 
    else:
        data['doc_name'] = None

    # 5. FIND DATES
    numeric_pattern = r'\d{2}[-/]\d{2}[-/]\d{4}'
    text_month_pattern = r'\d{1,2}[-\s][A-Za-z]{3}[-\s]\d{4}'
    date_pattern = f'(?:{numeric_pattern}|{text_month_pattern})'
    
    reg_match = re.search(r'(?:Registration|Regn|Reg\.)\s*Date[:\s]*(' + date_pattern + ')', text, re.IGNORECASE)
    data['reg_date_specific'] = reg_match.group(1) if reg_match else None

    rec_match = re.search(r'Receipt\s*date[:\s]*(' + date_pattern + ')', text, re.IGNORECASE)
    data['receipt_date_specific'] = rec_match.group(1) if rec_match else None

    if not data['reg_date_specific'] and not data['receipt_date_specific']:
        any_date = re.search(date_pattern, text)
        data['fallback_date'] = any_date.group(0) if any_date else None
    else:
        data['fallback_date'] = None

    return data

def analyze_row(row, doc_data):
    if not doc_data.get('doc_chassis'):
        return "No document found", "Ineligible", "MISSING DOCUMENT"

    chassis_match = str(row['Chassis number']).strip() == str(doc_data.get('doc_chassis')).strip()
    name_is_match = check_name_match(row['Customer Name'], doc_data.get('doc_name'))
    is_permanent = doc_data['reg_type'] == "Permanent"

    if chassis_match and name_is_match and is_permanent:
        return "Approved", "Approve", "None"

    if chassis_match and name_is_match and not is_permanent:
        return ("Uploaded document is temporary registration. Kindly upload VAHAN copy.", 
                "Hold", "TEMP REGISTRATION")

    if chassis_match and is_permanent and not name_is_match:
        found_name = doc_data.get('doc_name', 'Unknown')
        return (f"Name mismatch. Found: {found_name}. Provide relationship proof.", 
                "Hold", "NAME MISMATCH")

    if not chassis_match:
         return "Chassis Number mismatch", "Reject", "CHASSIS MISMATCH"

    return "Verification Failed", "Reject", "UNKNOWN ERROR"

def create_colored_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Verification')
    
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    fills = {
        "Reject": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Ineligible": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Hold": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Approve": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    }

    header = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    status_col_idx = header.get('RTO status')

    if status_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col_idx)
            val = str(cell.value).strip()
            if val in fills:
                cell.fill = fills[val]
    
    output_final = io.BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    return output_final

# --- STREAMLIT UI ---

st.title("🚗 Auto-Verification System (OCR Enabled)")
st.info("This system automatically uses OCR for scanned/non-searchable PDFs.")

col1, col2 = st.columns(2)

with col1:
    uploaded_excel = st.file_uploader("1. Upload Master Excel", type=["xlsx", "xls"])

with col2:
    uploaded_pdfs = st.file_uploader("2. Upload PDF Documents", type=["pdf"], accept_multiple_files=True)

if st.button("🚀 Run Verification"):
    if uploaded_excel and uploaded_pdfs:
        with st.spinner("Extracting data (OCR may take longer for scanned docs)..."):
            extracted_docs = []
            progress_bar = st.progress(0)
            
            for i, pdf_file in enumerate(uploaded_pdfs):
                text = extract_text_from_pdf_upload(pdf_file)
                doc_info = parse_document_data(text)
                if doc_info['doc_chassis']:
                    extracted_docs.append(doc_info)
                progress_bar.progress((i + 1) / len(uploaded_pdfs))
            
            df_docs = pd.DataFrame(extracted_docs)

            try:
                df_user = pd.read_excel(uploaded_excel)
                df_user.columns = df_user.columns.str.strip()
                
                # Validation
                required = ['Chassis number', 'Customer Name']
                if not all(col in df_user.columns for col in required):
                    st.error(f"Excel must contain: {required}")
                    st.stop()

                if not df_docs.empty:
                    merged_df = pd.merge(df_user, df_docs, left_on='Chassis number', right_on='doc_chassis', how='left')
                else:
                    merged_df = df_user.copy()
                    merged_df['doc_chassis'] = None

                results = []
                for _, row in merged_df.iterrows():
                    doc_data = {
                        'doc_name': row.get('doc_name'),
                        'doc_chassis': row.get('doc_chassis'),
                        'reg_type': row.get('reg_type', "Temporary"),
                        'vehicle_no': row.get('vehicle_no', "Not Found")
                    }
                    
                    final_date = row.get('reg_date_specific') or row.get('receipt_date_specific') or row.get('fallback_date')
                    remark, status, error_type = analyze_row(row, doc_data)
                    
                    results.append({
                        'Chassis number': row['Chassis number'],
                        'Customer name': row['Customer Name'],
                        'Registration date': final_date,
                        'Vehicle Num': doc_data['vehicle_no'],
                        'Certificate Attached': 'Yes' if doc_data['doc_chassis'] else 'No',
                        'RTO status': status,
                        'Specific Error': error_type,
                        'Remarks': remark
                    })

                final_df = pd.DataFrame(results)
                st.dataframe(final_df)

                processed_excel = create_colored_excel(final_df)
                st.download_button("📥 Download Report", processed_excel, "report.xlsx")

            except Exception as e:
                st.error(f"Error: {e}")
    else:
        st.error("Please upload both Excel and PDF files.")
